import sys
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem, QPushButton, QFileDialog
from PyQt5.QtCore import Qt
from openpyxl import Workbook, load_workbook

class ExcelApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle("Excel-like App")
        self.setGeometry(100, 100, 800, 600)

        self.layout = QVBoxLayout()

        # Table widget to display Excel-like grid
        self.table = QTableWidget(self)
        self.table.setRowCount(10)  # Default rows
        self.table.setColumnCount(5)  # Default columns
        self.table.setHorizontalHeaderLabels(["Column 1", "Column 2", "Column 3", "Column 4", "Column 5"])

        # Buttons to add/remove columns and rows dynamically
        self.add_column_button = QPushButton("Add Column", self)
        self.add_column_button.clicked.connect(self.add_column)
        self.layout.addWidget(self.add_column_button)

        self.remove_column_button = QPushButton("Remove Column", self)
        self.remove_column_button.clicked.connect(self.remove_column)
        self.layout.addWidget(self.remove_column_button)

        self.add_row_button = QPushButton("Add Row", self)
        self.add_row_button.clicked.connect(self.add_row)
        self.layout.addWidget(self.add_row_button)

        self.remove_row_button = QPushButton("Remove Row", self)
        self.remove_row_button.clicked.connect(self.remove_row)
        self.layout.addWidget(self.remove_row_button)

        # Buttons for file operations
        self.save_button = QPushButton("Save Excel File", self)
        self.save_button.clicked.connect(self.save_excel)
        self.layout.addWidget(self.save_button)

        self.load_button = QPushButton("Load Excel File", self)
        self.load_button.clicked.connect(self.load_excel)
        self.layout.addWidget(self.load_button)

        self.layout.addWidget(self.table)

        self.setLayout(self.layout)
        self.show()

    def add_column(self):
        # Add a new column to the table
        current_column_count = self.table.columnCount()
        self.table.setColumnCount(current_column_count + 1)

        # Set the header for the new column
        self.table.setHorizontalHeaderLabels([f"Column {i + 1}" for i in range(self.table.columnCount())])

    def remove_column(self):
        # Remove the last column
        current_column_count = self.table.columnCount()
        if current_column_count > 1:  # Prevent removing all columns
            self.table.setColumnCount(current_column_count - 1)

            # Update header names
            self.table.setHorizontalHeaderLabels([f"Column {i + 1}" for i in range(self.table.columnCount())])

    def add_row(self):
        # Add a new row to the table
        current_row_count = self.table.rowCount()
        self.table.setRowCount(current_row_count + 1)

    def remove_row(self):
        # Remove the last row
        current_row_count = self.table.rowCount()
        if current_row_count > 1:  # Prevent removing all rows
            self.table.setRowCount(current_row_count - 1)

    def save_excel(self):
        file_name, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx);;All Files (*)")
        if file_name:
            workbook = Workbook()
            sheet = workbook.active

            # Saving data to the Excel file
            for row in range(self.table.rowCount()):
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    if item is not None:
                        sheet.cell(row=row + 1, column=col + 1, value=item.text())
                    else:
                        sheet.cell(row=row + 1, column=col + 1, value="")

            workbook.save(file_name)
            print(f"File saved to {file_name}")

    def load_excel(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx);;All Files (*)")
        if file_name:
            workbook = load_workbook(file_name)
            sheet = workbook.active

            # Set the number of columns and rows dynamically based on the loaded Excel file
            self.table.setRowCount(sheet.max_row)
            self.table.setColumnCount(sheet.max_column)

            # Populating the table with data from the Excel file
            for row in range(sheet.max_row):
                for col in range(sheet.max_column):
                    value = sheet.cell(row=row + 1, column=col + 1).value
                    item = QTableWidgetItem(str(value) if value is not None else "")
                    self.table.setItem(row, col, item)

            print(f"File loaded from {file_name}")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = ExcelApp()
    sys.exit(app.exec_())
